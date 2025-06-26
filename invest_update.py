# dart_investment.py

import argparse
import os
import io
import re
from datetime import date, datetime
from zipfile import ZipFile, BadZipFile

import requests
import pandas as pd
from bs4 import BeautifulSoup, element
from dotenv import load_dotenv
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from pandas.api.types import is_integer_dtype, is_float_dtype

load_dotenv()
API_KEY = os.getenv("DART_API_KEY")
HEADERS = {"User-Agent": "Mozilla/5.0"}

def fetch_sales(session, bgn_de: str, end_de: str) -> pd.DataFrame:
    url = 'https://opendart.fss.or.kr/api/list.json'
    params = {
        'crtfc_key': API_KEY,
        'bgn_de': bgn_de,
        'end_de': end_de,
        'pblntf_detail_ty': 'I001',
        'page_count': 100,
        'last_reprt_at': 'Y',
    }
    all_reports = []
    resp = session.get(url, params={**params, 'page_no': 1}, timeout=10)
    resp.raise_for_status()
    data = resp.json()
    total_page = int(data.get('total_page', 1))
    all_reports.extend(data.get('list', []))
    for page in range(2, total_page + 1):
        resp = session.get(url, params={**params, 'page_no': page}, timeout=10)
        resp.raise_for_status()
        rpt = resp.json().get('list', [])
        if not rpt:
            break
        all_reports.extend(rpt)
    df = pd.DataFrame(all_reports)
    if df.empty or 'report_nm' not in df:
        return pd.DataFrame()
    mask = df['report_nm'].str.contains('신규시설', na=False)
    mask &= ~df['report_nm'].str.contains('자회사|철회', na=False)
    return df[mask].reset_index(drop=True)

def parse_investment_with_helpers(html) -> dict:
    table = html if isinstance(html, element.Tag) else \
            BeautifulSoup(html, 'html.parser').find(id="XFormD1_Form0_Table0")
    if table is None:
        return {}
    def get_val(keys):
        for tr in table.find_all('tr'):
            tds = tr.find_all('td')
            if len(tds) < 2:
                continue
            label = re.sub(r'\s+', '', tds[0].get_text() + (tds[1].get_text() if len(tds)>2 else ''))
            val   = tds[-1].get_text(strip=True)
            for k in sorted(keys, key=len, reverse=True):
                if k.replace('ㆍ','') in label or k in label:
                    return val
        return None
    def get_int(keys):
        v = get_val(keys)
        return None if not v or v=='-' else int(v.replace(',',''))
    def get_float(keys):
        v = get_val(keys)
        return None if not v or v=='-' else float(v.replace(',',''))
    def get_date(keys):
        v = get_val(keys)
        try:
            return datetime.strptime(v, '%Y-%m-%d').date() if v else None
        except:
            return None

    amt   = get_int(['투자금액(원)','투자금액'])
    equity= get_int(['자기자본(원)','자기자본'])
    return {
        '투자구분':    get_val(['투자대상']) or get_val(['투자구분']),
        '투자금액(백만원)': amt/1_000_000 if amt else None,
        '자기자본(백만원)': equity/1_000_000 if equity else None,
        '자기자본대비(%)': get_float(['자기자본대비(%)']),
        '결정일':      get_date(['이사회결의일(결정일)','결정일']),
        '시작일':      get_date(['시작일']),
        '종료일':      get_date(['종료일']),
    }

def parse_contract(session, rcept_no: str) -> dict:
    resp = session.get(
        'https://opendart.fss.or.kr/api/document.xml',
        params={'crtfc_key': API_KEY, 'rcept_no': rcept_no},
        timeout=10
    )
    resp.raise_for_status()
    try:
        z = ZipFile(io.BytesIO(resp.content))
    except BadZipFile:
        return {}
    fname = next(f for f in z.namelist() if f.lower().endswith(('.xml','.html')))
    html  = z.read(fname).decode('utf-8', 'ignore')
    soup  = BeautifulSoup(html, 'lxml')
    for tbl in soup.find_all('table'):
        if '투자구분' in tbl.get_text():
            return parse_investment_with_helpers(tbl)
    return {}

def fetch_history(code: str, page: int = 1) -> pd.DataFrame:
    url = f"https://finance.naver.com/item/sise_day.naver?code={code}&page={page}"
    resp = requests.get(url, headers=HEADERS)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, 'html.parser')
    recs = []
    for row in soup.select("table.type2 tr"):
        cols = row.find_all('td')
        if len(cols) < 7:
            continue
        d = cols[0].get_text(strip=True)
        c = cols[1].get_text(strip=True).replace(',', '')
        if not d or not c:
            continue
        try:
            dt = datetime.strptime(d, "%Y.%m.%d").date()
            cl = int(c)
        except:
            continue
        recs.append({'date': dt, 'close': cl})
    return pd.DataFrame(recs)

def fetch_closes(session, stock_code: str, rcept_dt: str):
    target_date = datetime.strptime(rcept_dt, "%Y%m%d").date()
    url = f"https://finance.naver.com/item/sise_day.naver?code={stock_code}&page=1"
    resp = session.get(url, headers=HEADERS, timeout=5)
    resp.raise_for_status()

    records = []
    soup = BeautifulSoup(resp.text, "lxml")
    for tr in soup.select("table.type2 tr"):
        cols = tr.find_all("td")
        if len(cols) != 7:
            continue
        date_txt = cols[0].get_text(strip=True)
        close_txt = cols[1].get_text(strip=True).replace(",", "")
        if not date_txt or not close_txt:
            continue
        try:
            dt = datetime.strptime(date_txt, "%Y.%m.%d").date()
            cl = int(close_txt)
            records.append((dt, cl))
        except:
            continue

    records.sort(key=lambda x: x[0], reverse=True)
    dates  = [dt for dt, _ in records]
    closes = [cl for _, cl in records]

    try:
        idx = dates.index(target_date)
        prev_close = closes[idx+1] if idx+1 < len(closes) else None
        today_close = closes[idx]
        next_close = closes[idx-1] if idx-1 >= 0 else None
    except (ValueError, IndexError):
        prev_close = today_close = next_close = None

    return {
        '전일종가': prev_close,
        '당일종가': today_close,
        '익일종가': next_close
    }

def update_excel(result_df: pd.DataFrame, excel_path: str, sheet_name: str='신규투자'):
    if os.path.exists(excel_path):
        try:
            wb = load_workbook(excel_path)
        except BadZipFile:
            wb = Workbook()
    else:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        existing = pd.read_excel(
            excel_path, sheet_name=sheet_name,
            parse_dates=['공시일','시작일','종료일'],
            usecols=list(result_df.columns)
        )
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(list(result_df.columns))
        existing = pd.DataFrame(columns=result_df.columns)

    header = [c.value for c in ws[1]]
    if existing.empty:
        new_rows = result_df
    else:
        mask = ~result_df.apply(
            lambda r: ((existing['공시회사']==r['공시회사']) &
                       (existing['공시일']==r['공시일'])).any(),
            axis=1
        )
        new_rows = result_df[mask]

    num_cols  = [c for c in result_df.columns
                 if is_integer_dtype(result_df[c]) or is_float_dtype(result_df[c])]
    col_index = {col: i+1 for i, col in enumerate(header)}
    align     = Alignment('center','center')
    font      = Font(size=10)

    for _, r in new_rows.iterrows():
        ws.append([r.get(col,'') for col in header])
        row = ws.max_row
        for col, idx in col_index.items():
            cell = ws.cell(row=row, column=idx)
            cell.alignment = align
            cell.font = font
            if col in ['공시일','시작일','종료일'] and pd.notna(r[col]):
                cell.number_format = 'yyyy-mm-dd'
            elif col == '자기자본대비(%)' and pd.notna(r[col]):
                cell.number_format = '#,##0.00'
            elif col in num_cols:
                cell.number_format = '#,##0'

    wb.save(excel_path)

def filter_new_rows(result_df: pd.DataFrame, excel_path: str,
                    sheet_name: str='신규투자') -> pd.DataFrame:
    if not os.path.exists(excel_path):
        return result_df.copy()

    existing = pd.read_excel(
        excel_path,
        sheet_name=sheet_name,
        usecols=['공시회사','공시일','투자구분'],
    )
    existing['공시일'] = pd.to_datetime(existing['공시일'], errors='coerce').dt.date
    existing['공시회사'] = existing['공시회사'].astype(str).str.strip()
    existing['투자구분'] = existing['투자구분'].astype(str).str.strip()

    existing_keys = set(zip(
        existing['공시회사'],
        existing['공시일'],
        existing['투자구분']
    ))

    def is_new(row):
        key = (
            row['공시회사'].strip() if isinstance(row['공시회사'], str) else row['공시회사'],
            row['공시일'],
            row['투자구분'].strip() if isinstance(row['투자구분'], str) else row['투자구분']
        )
        return key not in existing_keys

    return result_df[result_df.apply(is_new, axis=1)].copy()

def fill_next_close(session, excel_path: str, sheet_name: str='신규투자'):
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]

    header = [c.value for c in ws[1]]
    idx_code  = header.index('종목코드')   + 1
    idx_rcept = header.index('공시일')     + 1
    idx_prev  = header.index('전일종가')   + 1
    idx_today = header.index('당일종가')   + 1
    idx_next  = header.index('익일종가')   + 1

    for row in range(2, ws.max_row + 1):
        if ws.cell(row, idx_next).value is not None:
            continue

        code = ws.cell(row, idx_code).value
        rcept_cell = ws.cell(row, idx_rcept).value

        if rcept_cell is None:
            continue

        if hasattr(rcept_cell, 'strftime'):
            rcept_dt = rcept_cell.strftime('%Y%m%d')
        else:
            try:
                rcept_dt = datetime.strptime(str(rcept_cell), '%Y-%m-%d').strftime('%Y%m%d')
            except ValueError:
                continue

        closes = fetch_closes(session, code, rcept_dt)

        for idx, key in ((idx_prev, '전일종가'),
                         (idx_today, '당일종가'),
                         (idx_next, '익일종가')):
            cell = ws.cell(row, idx)
            cell.value = closes[key]
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='center', vertical='center')

    print('✅ 익일 종가 업데이트 완료')
    wb.save(excel_path)


def main():
    parser = argparse.ArgumentParser(
        description="DART 신규시설 공시를 조회해서 엑셀로 저장합니다."
    )
    parser.add_argument(
        "--date", "-d",
        type=str,
        default=date.today().strftime("%Y%m%d"),
        help="조회 기준일자 (YYYYMMDD). 기본: 오늘"
    )
    args = parser.parse_args()
    target = args.date

    sess = requests.Session()
    sales = fetch_sales(sess, bgn_de=target, end_de=target)

    fill_next_close(sess, '국내 주요 공시 정리.xlsx')
    if sales.empty:
        print("오늘 신규시설 투자 공시가 없습니다.")
        return 
    
    parsed = []
    for rec in tqdm(sales.to_dict("records")):
        d = parse_contract(sess, str(rec["rcept_no"])) or {}
        d.update({
            "공시회사": rec["corp_name"],
            "공시일":   datetime.strptime(target, "%Y%m%d").date(),
            "종목코드": rec["stock_code"]
        })
        d.update(fetch_closes(sess, rec["stock_code"], target))
        parsed.append(d)

    final_df = pd.DataFrame(parsed, columns=[
        "공시회사","공시일","종목코드","투자구분","투자금액(백만원)",
        "자기자본(백만원)","자기자본대비(%)","결정일","시작일","종료일",
        "전일종가","당일종가","익일종가"
    ]).sort_values("공시일")

    new_df = filter_new_rows(final_df, '국내 주요 공시 정리.xlsx')

    if new_df is None or new_df.empty:
        print("업데이트할 공시가 없습니다.")

    update_excel(new_df, "국내 주요 공시 정리.xlsx")

if __name__ == "__main__":
    main()