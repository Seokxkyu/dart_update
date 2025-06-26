import re
import io
import zipfile
import requests
import pandas as pd
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from dotenv import load_dotenv 
import os

load_dotenv()
API_KEY = os.getenv("DART_API_KEY")
EXCEL_PATH = '국내 주요 공시 정리.xlsx'
SHEET_NAME = '합병'

def clean_number(val: str) -> int | None:
    if not isinstance(val, str):
        return None
    s = re.sub(r'\(.*?\)', '', val)
    digits = re.sub(r'[^\d]', '', s)
    return int(digits) if digits else None


def parse_amount(val: str) -> int | None:
    if not isinstance(val, str):
        return None
    s = re.sub(r'\(.*?\)', '', val)
    digits = re.sub(r'[^\d]', '', s)
    if not digits:
        return None
    base = int(digits)
    if '백만' in s or '백' in s:
        return base * 1_000_000
    return base

def parse_merger_overview(rcept_no: str, corp_name: str) -> dict | None:
    url  = 'https://opendart.fss.or.kr/api/document.xml'
    resp = requests.get(url, params={'crtfc_key': API_KEY, 'rcept_no': rcept_no})
    resp.raise_for_status()
    with zipfile.ZipFile(io.BytesIO(resp.content)) as z:
        with z.open(z.namelist()[0]) as f:
            xml = f.read().decode('utf-8', errors='replace')

    part = xml.split("(1) 합병 당사회사의 개요", 1)
    if len(part) < 2:
        return None
    block = "(1) 합병 당사회사의 개요" + part[1]
    soup  = BeautifulSoup(block, 'html.parser')
    table = soup.find('table')
    rows  = table.find_all('tr') if table else []

    merger_data = {}
    mapping    = {
        r'법인명': 'corp_name',
        r'납입자본금': 'capital',
        r'자산총액': 'assets',
        r'주권상장여부': 'listing',
        r'발행주식.*수': 'shares'
    }
    for tr in rows:
        texts = [td.get_text(strip=True) for td in tr.find_all('td')]
        if len(texts) != 3:
            continue
        for pat, field in mapping.items():
            if re.search(pat, texts[0]):
                merger_data[field] = {'합병법인': texts[1], '피합병법인': texts[2]}
                break

    starts = [m.start() for m in re.finditer(re.escape("1. 사업의 개요"), xml)]
    blocks = [
        xml[starts[i]: (starts[i+1] if i+1 < len(starts) else len(xml))]
        for i in range(len(starts))
    ]
    end_marker = re.compile(r"(?:1\. 합병의 개요|2\. 주요 제품 및 서비스)")
    combined = []
    for blk in blocks:
        if "나. 회사의 현황" in blk:
            sub   = blk.split("나. 회사의 현황", 1)[1]
            soup2 = BeautifulSoup(sub, 'html.parser')
            para  = next((p.get_text(strip=True) for p in soup2.find_all('p') if p.get_text(strip=True)), "")
        else:
            m     = end_marker.search(blk)
            cut   = blk[:m.start()] if m else blk
            soup2 = BeautifulSoup(cut, 'html.parser')
            para  = next((p.get_text(strip=True) for p in soup2.find_all('p') if p.get_text(strip=True)), "")
        combined.append(para)

    biz_merge  = combined[0] if combined else ""
    biz_target = combined[1] if len(combined)>1 else ""

    def G(d,f,r): return d.get(f,{}).get(r)
    def N(d,f,r): return parse_amount(G(d,f,r))

    return {
        '공시회사': corp_name,
        '합병법인':     G(merger_data,'corp_name','합병법인'),
        '피합병법인':   G(merger_data,'corp_name','피합병법인'),
        '납입자본금(합병)':   N(merger_data,'capital','합병법인'),
        '납입자본금(피합병)': N(merger_data,'capital','피합병법인'),
        '자산총액(합병)':     N(merger_data,'assets','합병법인'),
        '자산총액(피합병)':   N(merger_data,'assets','피합병법인'),
        '합병법인 상장':      G(merger_data,'listing','합병법인'),
        '피합병법인 상장':    G(merger_data,'listing','피합병법인'),
        '발행주식수(합병)':   clean_number(G(merger_data,'shares','합병법인')),
        '발행주식수(피합병)': clean_number(G(merger_data,'shares','피합병법인')),
        '사업개요(합병)':     biz_merge,
        '사업개요(피합병)':   biz_target
    }

def get_merger_reports_for_date(date_str: str) -> pd.DataFrame:
    url    = 'https://opendart.fss.or.kr/api/list.json'
    params = {
        'crtfc_key':       API_KEY,
        'pblntf_detail_ty':'C004',
        'bgn_de':          date_str,
        'end_de':          date_str,
        'last_reprt_at':   'Y',
        'page_no':         1,
        'page_count':      100
    }
    resp = requests.get(url, params=params)
    data = resp.json().get('list', [])
    df   = pd.DataFrame(data)
    df   = df[df['report_nm'].str.contains(r"증권신고서\(합병", na=False)]

    recs = []
    for _,row in df.iterrows():
        mv = parse_merger_overview(row['rcept_no'], row['corp_name'])
        if mv:
            mv['최종보고일'] = row['rcept_dt']
            recs.append(mv)

    out = pd.DataFrame(recs)

    if out.empty:
        return out
    
    out['사업개요'] = out.apply(
        lambda x: x['사업개요(피합병)'] if re.search(r"인수목적|스팩|SPAC", x['합병법인']) else x['사업개요(합병)'],
        axis=1
    )
    out = out.drop(columns=['사업개요(합병)','사업개요(피합병)'])

    cols = ['공시회사','합병법인','피합병법인','최종보고일',
            '납입자본금(합병)','납입자본금(피합병)',
            '자산총액(합병)','자산총액(피합병)',
            '합병법인 상장','피합병법인 상장',
            '발행주식수(합병)','발행주식수(피합병)',
            '사업개요']
    return out[cols]

def update_excel(df_all: pd.DataFrame):
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    header = [c.value for c in ws[1]]
    col_map = {col:i for i,col in enumerate(df_all.columns)}

    if '최초보고일' not in header:
        idx = header.index('최종보고일') + 1
        ws.insert_cols(idx+1)
        ws.cell(row=1, column=idx+1, value='최초보고일')
        header.insert(idx, '최초보고일')

    existing = {}
    idx_corp  = header.index('합병법인')
    idx_target= header.index('피합병법인')
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        key = (row[idx_corp].value, row[idx_target].value)
        existing[key] = row[0].row

    date_cols  = {'최종보고일','최초보고일'}
    comma_cols = {'납입자본금(합병)','납입자본금(피합병)','자산총액(합병)','자산총액(피합병)','발행주식수(합병)','발행주식수(피합병)'}
    wrap_col   = '사업개요'

    def apply_fmt(cell, col):
        if col in date_cols and isinstance(cell.value, str):
            try:
                cell.value = datetime.strptime(cell.value, '%Y%m%d')
                cell.number_format = 'yyyy-mm-dd'
            except:
                pass
        elif col in comma_cols and isinstance(cell.value, (int,float)):
            cell.number_format = '#,##0'
        if col==wrap_col:
            cell.alignment=Alignment(wrap_text=True)
        else:
            cell.alignment=Alignment(horizontal='center',vertical='center')
        cell.font=Font(size=10)

    start_row = ws.max_row + 1
    for idx, series in df_all.iterrows():
        key = (series["합병법인"], series["피합병법인"])
        if key in existing:
            r = existing[key]
            for col_idx, col_name in enumerate(header, start=1):
                if col_name in df_all.columns:
                    val = series[col_name]
                    cell = ws.cell(row=r, column=col_idx)
                    if cell.value != val:
                        cell.value = val
                        apply_fmt(cell, col_name)
        else:
            r = start_row
            for col_idx, col_name in enumerate(header, start=1):
                if col_name in df_all.columns:
                    val = series[col_name]
                elif col_name == "최초보고일":
                    val = series["최종보고일"]
                else:
                    continue
                cell = ws.cell(row=r, column=col_idx, value=val)
                apply_fmt(cell, col_name)
            start_row += 1

    wb.save(EXCEL_PATH)
    print("✅ 업데이트 완료료")


def main(date_str: str):
    df_new = get_merger_reports_for_date(date_str)
    if not df_new.empty:
        update_excel(df_new)
    else:
        
        print(f"오늘 합병 관련 증권신고가 없습니다.")
        return

if __name__=='__main__':
    import argparse
    parser = argparse.ArgumentParser(description="증권보고서 업데이트")
    parser.add_argument(
        "--date",
        type=str,
        default=datetime.now().strftime("%Y%m%d"),
        help="조회할 날짜(YYYYMMDD), 기본값은 오늘"
    )
    args = parser.parse_args()
    main(args.date)