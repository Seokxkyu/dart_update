import os
import io
import re
import argparse
from datetime import datetime
import requests
import pandas as pd
from zipfile import ZipFile, BadZipFile
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from pandas.api.types import (
    is_integer_dtype,
    is_float_dtype
)
from bs4 import BeautifulSoup
from dotenv import load_dotenv         

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/114.0.0.0 Safari/537.36"
    )
}
load_dotenv()
API_KEY = os.getenv("DART_API_KEY")

def fetch_sales(session, target_date: str) -> pd.DataFrame:
    url = 'https://opendart.fss.or.kr/api/list.json'
    base_params = {
        'crtfc_key': API_KEY,
        'bgn_de': target_date,
        'end_de': target_date,
        'pblntf_detail_ty': 'I001',
        'page_count': 100,
        'last_reprt_at': 'Y',
    }

    all_reports = []
    resp = session.get(url, params={**base_params, 'page_no': 1}, timeout=10)
    resp.raise_for_status()
    data = resp.json()
    total_page = int(data.get('total_page', 1))
    all_reports.extend(data.get('list', []))

    for page in range(2, total_page + 1):
        resp = session.get(url, params={**base_params, 'page_no': page}, timeout=10)
        resp.raise_for_status()
        reports = resp.json().get('list', [])
        if not reports:
            break
        all_reports.extend(reports)

    df = pd.DataFrame(all_reports)

    if df.empty or 'report_nm' not in df.columns:
        return pd.DataFrame()

    return df[
        df['report_nm'].str.contains('단일판매') &
        ~df['report_nm'].str.contains('정정|해지')
    ].reset_index(drop=True)


def parse_contract(session, rcept_no: str) -> dict:
    resp = session.get(
        'https://opendart.fss.or.kr/api/document.xml',
        params={'crtfc_key': API_KEY, 'rcept_no': rcept_no},
        timeout=10
    )
    resp.raise_for_status()
    try:
        z = ZipFile(io.BytesIO(resp.content))
    except BadZipFile:
        return {}
    fname = next(f for f in z.namelist() if f.lower().endswith(('.xml', '.html')))
    html = z.read(fname).decode('utf-8', errors='ignore')
    soup = BeautifulSoup(html, 'lxml')

    contract_table = None
    for tbl in soup.find_all('table'):
        text = tbl.get_text()
        if any(k in text for k in ['계약금액', '판매ㆍ공급계약', '세부내용', '계약내역']):
            contract_table = tbl
            break
    if contract_table is None:
        return {}

    def get_val(keys):
        for tr in contract_table.find_all('tr'):
            tds = tr.find_all('td')
            if len(tds) < 2:
                continue
            label = re.sub(r'\s+', '', tds[0].get_text() + tds[1].get_text())
            if any(k.replace('ㆍ', '') in label or k in label for k in keys):
                return tds[-1].get_text(strip=True).replace(',', '')
        return None

    def get_int(keys):
        raw = get_val(keys)
        if raw in (None, '', '-'):
            return None
        return int(raw)

    def get_float(keys):
        raw = get_val(keys)
        if raw in (None, '', '-'):
            return None
        return float(raw)

    name = get_val([
        '체결계약명', '판매ㆍ공급계약내용', '판매ㆍ공급계약 내용',
        '판매ㆍ공급계약 구분', '공급계약 구분', '세부내용', '공급계약내용'
    ])
    if not name:
        name = get_val(['계약내역'])

    raw_start = get_val(['시작일'])
    if raw_start in (None, '', '-'):
        raw_start = get_val(['계약(수주)일자', '계약(수주)일'])

    return {
        '내용':             name or '',
        '계약 금액(억)':     int(get_int(['계약금액', '계약금액총액']) or 0) / 100_000_000,
        '매출액 대비(%) (A)': float(get_float(['매출액대비', '매출액대비(%)']) or 0),
        '시작일 (s)':        raw_start,
        '종료일 (e)':        get_val(['종료일']),
        '계약상대':          get_val(['계약상대', '계약상대방']),
    }

def fetch_market_info(session, stock_code: str) -> dict:
    url = f"https://navercomp.wisereport.co.kr/v2/company/c1010001.aspx?cmp_cd={stock_code}&cn="
    res = session.get(url, timeout=10)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, 'lxml')

    wics = None
    for dt in soup.select("td.td0101 dl dt.line-left"):
        txt = dt.get_text(strip=True)
        if txt.startswith("WICS"):
            wics = txt.split(":", 1)[1].strip()
            break

    mktcap = None
    tbl = soup.find("table", id="cTB11")
    if tbl:
        for tr in tbl.find_all("tr"):
            th = tr.find("th", class_="txt")
            if th and "시가총액" in th.get_text():
                raw = tr.find("td", class_="num").get_text(strip=True)
                num_str = raw.replace("억원", "").replace(",", "")
                try:
                    mktcap = int(num_str)
                except ValueError:
                    mktcap = None
                break

    return {'업종 분류': wics, '시가총액(억)': mktcap}

def filter_new_rows(result_df: pd.DataFrame, existing_df: pd.DataFrame) -> pd.DataFrame:
    if existing_df.empty:
        return result_df.copy()
    key_cols = ['날짜 (D)', '공시회사', '계약 금액(억)']
    existing_keys = set(
        zip(
            existing_df['날짜 (D)'].dt.strftime('%Y%m%d'),
            existing_df['공시회사'],
            existing_df['계약 금액(억)']
        )
    )
    def is_new(row):
        return (
            row['날짜 (D)'].strftime('%Y%m%d'),
            row['공시회사'],
            row['계약 금액(억)']
        ) not in existing_keys
    return result_df[result_df.apply(is_new, axis=1)]


def update_excel(result_df: pd.DataFrame, excel_path: str):
    if os.path.exists(excel_path):
        try:
            wb = load_workbook(excel_path)
        except BadZipFile:
            wb = Workbook()
    else:
        wb = Workbook()

    if 'main' in wb.sheetnames:
        ws = wb['main']
        existing_df = pd.read_excel(
            excel_path,
            sheet_name='main',
            usecols=['날짜 (D)', '공시회사', '계약 금액(억)', 'Cnt'],
            parse_dates=['날짜 (D)']
        )
    else:
        ws = wb.create_sheet('main')
        existing_df = pd.DataFrame(columns=list(result_df.columns) + ['Cnt'])
        ws.append(list(result_df.columns) + ['Cnt'])

    new_rows = filter_new_rows(result_df, existing_df)

    if not existing_df.empty:
        existing_df_sorted = existing_df.sort_values(['공시회사', '날짜 (D)'])
        existing_latest = existing_df_sorted.groupby('공시회사', as_index=False).last()
        numeric_cnt = pd.to_numeric(existing_latest['Cnt'], errors='coerce').fillna(0).astype(int)
        existing_max_cnt = dict(zip(
            existing_latest['공시회사'],
            numeric_cnt
        ))
    else:
        existing_max_cnt = {}

    header_row = [cell.value for cell in ws[1]]
    numeric_cols = [
        col for col in result_df.columns
        if is_integer_dtype(result_df[col]) or is_float_dtype(result_df[col])
    ]
    col_idx_map = {
        col: header_row.index(col) + 1
        for col in numeric_cols
        if col in header_row
    }

    next_cnt = {
        company: existing_max_cnt.get(company, 0) + 1
        for company in existing_max_cnt
    }

    for _, row in new_rows.iterrows():
        company = row['공시회사']
        if company not in next_cnt:
            next_cnt[company] = 1

        this_cnt = next_cnt[company]
        next_cnt[company] += 1

        row_values = []
        for col in header_row:
            if col == 'Cnt':
                row_values.append(this_cnt)
            else:
                row_values.append(row.get(col, ''))

        ws.append(row_values)
        new_row_idx = ws.max_row

        for idx, col in enumerate(header_row, start=1):
            cell = ws.cell(row=new_row_idx, column=idx)
            value = row.get(col, '')

            if col in ['날짜 (D)', '시작일 (s)', '종료일 (e)'] and pd.notna(value):
                cell.number_format = 'yyyy-mm-dd'
            elif col in col_idx_map:
                if is_integer_dtype(result_df[col]):
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '#,##0.00'
            if col not in ['내용', '계약상대']:
                cell.alignment = Alignment(horizontal='center')

    wb.save(excel_path)

# def update_missing_next_close(session, df_sales, excel_path):
#     wb = load_workbook(excel_path)
#     ws = wb['main']
#     hdr = [c.value for c in ws[1]]
#     idx_date       = hdr.index('날짜 (D)')    + 1
#     idx_name       = hdr.index('공시회사')    + 1
#     idx_next_close = hdr.index('익일종가(원)') + 1

#     code_map = {
#         (row['corp_name'], row['rcept_dt']): str(row['stock_code']).zfill(6)
#         for _, row in df_sales.iterrows()
#     }

#     for r in range(2, ws.max_row+1):
#         if ws.cell(r, idx_next_close).value is not None:
#             continue
#         raw_dt = ws.cell(r, idx_date).value
#         if hasattr(raw_dt, 'strftime'):
#             dt_str = raw_dt.strftime('%Y%m%d')
#         else:
#             dt_str = datetime.strptime(str(raw_dt), '%Y-%m-%d').strftime('%Y%m%d')
#         comp = ws.cell(r, idx_name).value
#         code = code_map.get((comp, dt_str))
#         if not code:
#             continue
#         _, _, nc = fetch_closes_simple(session, code, dt_str)
#         if nc is not None:
#             ws.cell(r, idx_next_close, nc)

#     wb.save(excel_path)

# def fetch_closes_simple(session, stock_code: str, rcept_dt: str):
#     target = datetime.strptime(rcept_dt, "%Y%m%d").strftime("%Y.%m.%d")

#     url = f"https://finance.naver.com/item/sise_day.naver?code={stock_code}&page=1"
#     resp = session.get(url, headers=HEADERS, timeout=5)
#     resp.raise_for_status()

#     soup = BeautifulSoup(resp.text, "lxml")
#     data = []
#     for tr in soup.select("table.type2 tr"):
#         cols = tr.find_all("td")
#         if len(cols) == 7:
#             date = cols[0].get_text(strip=True)
#             if date:
#                 close = int(cols[1].get_text(strip=True).replace(",", ""))
#                 data.append((date, close))

#     for idx, (dt, cl) in enumerate(data):
#         if dt == target:
#             prev_  = data[idx+1][1] if idx+1 < len(data) else None
#             today_ = cl
#             next_  = data[idx-1][1] if idx-1 >= 0 else None
#             return prev_, today_, next_
    
#     return None, None, None

def fetch_closes(session, stock_code: str, rcept_dt: str):
    target_date = datetime.strptime(rcept_dt, "%Y%m%d").date()
    url = f"https://finance.naver.com/item/sise_day.naver?code={stock_code}&page=1"
    resp = session.get(url, headers=HEADERS, timeout=5)
    resp.raise_for_status()
    records = []
    soup = BeautifulSoup(resp.text, "lxml")
    for tr in soup.select("table.type2 tr"):
        cols = tr.find_all("td")
        if len(cols) != 7:
            continue
        date_txt = cols[0].get_text(strip=True)
        close_txt = cols[1].get_text(strip=True).replace(",", "")
        if not date_txt or not close_txt:
            continue
        try:
            dt = datetime.strptime(date_txt, "%Y.%m.%d").date()
            cl = int(close_txt)
        except Exception:
            continue
        records.append((dt, cl))
    records.sort(key=lambda x: x[0], reverse=True)
    dates = [dt for dt, _ in records]
    closes = [cl for _, cl in records]
    try:
        idx = dates.index(target_date)
    except ValueError:
        return None, None, None
    prev_close = closes[idx + 1] if idx + 1 < len(closes) else None
    today_close = closes[idx]
    next_close = closes[idx - 1] if idx - 1 >= 0 else None
    return prev_close, today_close, next_close


def main(target_date: str, excel_path: str):
    session = requests.Session()
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/114.0.0.0 Safari/537.36"
        )
    })

    df = fetch_sales(session, target_date)
    if df.empty:
        print("오늘 단일판매 공시가 없습니다.")
        return

    mapping = {'Y': 'KS', 'K': 'KQ'}
    codes = df['stock_code'].astype(str).unique()

    market_infos = {code: fetch_market_info(session, code) for code in codes}

    records = []
    for _, row in df.iterrows():
        print(f"[DEBUG] 호출 → fetch_closes_simple(code={row['stock_code']}, date={row['rcept_dt']})")
        detail = parse_contract(session, row['rcept_no'])
        market = market_infos.get(str(row['stock_code']), {})
        if market.get('업종 분류') == '건설':
            continue

        prev_c, today_c, next_c = fetch_closes(
            session,
            str(row['stock_code']),
            row['rcept_dt']
        )
        print(f"[DEBUG] 반환 ← 전일:{prev_c}, 당일:{today_c}, 익일:{next_c}")

        records.append({
            '공시회사': row['corp_name'],
            '날짜 (D)': row['rcept_dt'],
            '거래소': mapping.get(row['corp_cls'], ''),
            '내용': detail.get('내용', ''),
            '계약 금액(억)': detail.get('계약 금액(억)', 0.0),
            '매출액 대비(%) (A)': detail.get('매출액 대비(%) (A)', 0.0),
            '계약상대': detail.get('계약상대', ''),
            '시작일 (s)': detail.get('시작일 (s)'),
            '종료일 (e)': detail.get('종료일 (e)'),
            '업종 분류': market.get('업종 분류', ''),
            '시가총액(억)': market.get('시가총액(억)', 0),
            '전일종가(원)': prev_c,
            '당일종가(원)': today_c,
            '익일종가(원)': next_c
        })

    result_df = pd.DataFrame(records)
    result_df['날짜 (D)'] = pd.to_datetime(result_df['날짜 (D)'], format='%Y%m%d', errors='coerce')
    result_df['시작일 (s)'] = pd.to_datetime(result_df['시작일 (s)'], format='%Y-%m-%d', errors='coerce')
    result_df['종료일 (e)'] = pd.to_datetime(result_df['종료일 (e)'], format='%Y-%m-%d', errors='coerce')
    print(result_df[['공시회사','전일종가(원)','당일종가(원)','익일종가(원)']])

    for col in ['계약 금액(억)', '매출액 대비(%) (A)', '시가총액(억)']:
        if result_df[col].dtype == object:
            result_df[col] = result_df[col].astype(str).str.replace(',', '').astype(float if '금액' in col or '%' in col else int)

    update_excel(result_df, excel_path)
    # update_missing_next_close(session, df, excel_path)
    print("✅ 공시 업데이트 완료")

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="DART 단일판매 공시 업데이트")
    parser.add_argument(
        "--date",
        type=str,
        default=datetime.now().strftime("%Y%m%d"),
        help="조회할 날짜(YYYYMMDD), 기본값은 오늘"
    )
    parser.add_argument(
        "--excel",
        type=str,
        default="국내 주요 공시 정리.xlsx",
        help="업데이트할 엑셀 파일 경로"
    )
    args = parser.parse_args()
    main(args.date, args.excel)